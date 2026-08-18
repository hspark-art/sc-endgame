#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""끝장전 기록실 사이트를 통째로 다시 만듭니다.

    python3 tools/build.py

입력   data/endgame.json  (정본 기록)
       data/videos.json   (경기 → 다시보기 영상 매핑)
출력   index.html, p/*.html, sheets.html, csv/*, xlsx/*

원본 수치는 손대지 않습니다. 여기서 만드는 값은 전부 원본에서 다시 계산한 것이라
언제 몇 번을 돌려도 같은 결과가 나옵니다.
"""

import csv
import io
import json
import os
import re
import sys
from datetime import datetime, timezone, timedelta

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)

import admin_php                                 # noqa: E402
import audit                                     # noqa: E402
import render                                    # noqa: E402
import stats                                     # noqa: E402
from slug import unique_slugs                    # noqa: E402

KST = timezone(timedelta(hours=9))


def read_json(path, default=None):
    if not os.path.exists(path):
        return default
    with io.open(path, encoding='utf-8') as f:
        return json.load(f)


def write(path, text):
    full = os.path.join(ROOT, path)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with io.open(full, 'w', encoding='utf-8', newline='\n') as f:
        f.write(text)
    return len(text.encode('utf-8'))


def write_csv(name, header, rows):
    """구글 시트·엑셀 양쪽에서 한글이 깨지지 않도록 BOM 을 붙여 저장합니다."""
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator='\n')
    w.writerow(header)
    w.writerows(rows)
    full = os.path.join(ROOT, 'csv', name)
    os.makedirs(os.path.dirname(full), exist_ok=True)
    with io.open(full, 'w', encoding='utf-8-sig', newline='') as f:
        f.write(buf.getvalue())
    return len(rows)


def pct_val(w, l):
    t = w + l
    return round(w / t * 100, 1) if t else ''


# ── 데이터 읽기 ────────────────────────────────────────────────
def load():
    data = read_json(os.path.join(ROOT, 'data', 'endgame.json'))
    if not data:
        raise SystemExit('data/endgame.json 이 없습니다.')
    videos = read_json(os.path.join(ROOT, 'data', 'videos.json'), {}) or {}
    vmap = videos.get('matches') or {}

    matched = 0
    for m in data['matches']:
        url = vmap.get(stats.match_key(m))
        m['youtubeUrl'] = url or None
        if url:
            matched += 1
    return data, matched


# ── 파생 통계 ──────────────────────────────────────────────────
def enrich(data):
    players = data['players']
    matches = data['matches']
    names = [p['name'] for p in players]

    slugs = unique_slugs(names)
    race_of = {p['name']: p['race'] for p in players}
    strk = stats.streaks(matches, names)
    p_year = stats.player_yearly(matches, names)
    m_year = stats.map_yearly(matches)
    yearly = stats.yearly(matches)
    years = [y['year'] for y in yearly]

    for p in players:
        p['slug'] = slugs[p['name']]
        p['streak'] = strk[p['name']]
        p['yearly'] = p_year.get(p['name'], {})
    for mp in data['maps']:
        mp['yearly'] = m_year.get(mp['name'], {})

    ctx = {
        'slugs': slugs, 'raceOf': race_of, 'streaks': strk,
        'playerYearly': p_year, 'years': years, 'yearly': yearly,
        'matches': matches, 'players': players,
        'mu': stats.global_matchups(matches),
        'records': stats.records(data, slugs, strk),
        'rivalries': stats.rivalries(players, 15),
        'mapCoveredSets': sum(mp['totalSets'] for mp in data['maps']),
    }
    return ctx


# ── 허브가 쓰는 데이터 (상대 전적 상세는 선수 페이지에만) ────────
def hub_data(data, ctx, built_at):
    players = [{
        'name': p['name'], 'slug': p['slug'], 'race': p['race'],
        'matchWin': p['matchWin'], 'matchLoss': p['matchLoss'],
        'setWin': p['setWin'], 'setLoss': p['setLoss'],
        'appearances': p['appearances'],
        'firstDate': p['firstDate'], 'lastDate': p['lastDate'],
        'streak': p['streak'], 'yearly': p['yearly'],
    } for p in data['players']]

    maps = [{
        'name': m['name'], 'totalSets': m['totalSets'], 'daysUsed': m['daysUsed'],
        'firstDate': m['firstDate'], 'lastDate': m['lastDate'],
        'matchup': m['matchup'], 'yearly': m['yearly'],
    } for m in data['maps']]

    return {
        'builtAt': built_at,
        'global': data['global'],
        'years': ctx['years'],
        'yearly': ctx['yearly'],
        'mu': ctx['mu'],
        'records': ctx['records'],
        'rivalries': ctx['rivalries'],
        'mapCoveredSets': ctx['mapCoveredSets'],
        'slugs': ctx['slugs'],
        'raceOf': ctx['raceOf'],
        'players': players,
        'maps': maps,
        'matches': data['matches'],
    }


# ── CSV ────────────────────────────────────────────────────────
CSV_SPEC = [
    ('players.csv', '선수', '선수별 통산 전적과 종족별 세트 전적'),
    ('matches.csv', '경기', '매치 한 줄씩 — 날짜·대진·스코어·맵·다시보기'),
    ('sets.csv', '세트별 맵', '매치 안의 세트마다 어떤 맵을 썼는지'),
    ('maps.csv', '맵', '맵별 사용 세트 수와 종족 상성'),
    ('headtohead.csv', '상대전적', '선수 대 선수 맞대결 기록'),
    ('yearly.csv', '연도별', '연도별 매치·세트 수와 종족 상성'),
]


def build_csv(data, ctx):
    counts = {}

    rows = []
    for p in sorted(data['players'], key=lambda x: (-x['matchWin'], x['name'])):
        v = p['vsRace']
        rows.append([
            p['name'], p['slug'], render.RACE_LABEL[p['race']], p['race'],
            p['matchWin'], p['matchLoss'], pct_val(p['matchWin'], p['matchLoss']),
            p['setWin'], p['setLoss'], pct_val(p['setWin'], p['setLoss']),
            p['appearances'], p['firstDate'], p['lastDate'],
            p['streak']['bestWin'], p['streak']['bestLoss'],
            v['T']['w'], v['T']['l'], v['P']['w'], v['P']['l'], v['Z']['w'], v['Z']['l'],
        ])
    counts['players.csv'] = write_csv('players.csv', [
        '선수', '슬러그', '종족', '종족코드', '매치승', '매치패', '매치승률',
        '세트승', '세트패', '세트승률', '출전', '첫출전', '최근출전',
        '최다연승', '최다연패',
        'vs테란_세트승', 'vs테란_세트패', 'vs프로토스_세트승', 'vs프로토스_세트패',
        'vs저그_세트승', 'vs저그_세트패',
    ], rows)

    rows, set_rows = [], []
    for m in data['matches']:
        a, b = m['players']
        maps = [x for x in (m.get('maps') or [])]
        rows.append([
            m['date'], a, m['race'][a], b, m['race'][b],
            m['setWins'][a], m['setWins'][b], m['winner'],
            m['setWins'][a] + m['setWins'][b],
            ' | '.join(x for x in maps if x), m.get('youtubeUrl') or '',
        ])
        for i, name in enumerate(maps, 1):
            set_rows.append([m['date'], a, b, i, name])
    counts['matches.csv'] = write_csv('matches.csv', [
        '날짜', '선수A', '종족A', '선수B', '종족B', '세트A', '세트B', '승자',
        '총세트', '맵목록', '다시보기',
    ], rows)
    counts['sets.csv'] = write_csv('sets.csv',
                                   ['날짜', '선수A', '선수B', '세트번호', '맵'], set_rows)

    rows = []
    for m in sorted(data['maps'], key=lambda x: -x['totalSets']):
        zp, tz, pt = m['matchup']['Z-P'], m['matchup']['T-Z'], m['matchup']['P-T']
        br = m['byRace']
        rows.append([
            m['name'], m['totalSets'], m['daysUsed'], m['firstDate'], m['lastDate'],
            zp['w'], zp['l'], pct_val(zp['w'], zp['l']),
            tz['w'], tz['l'], pct_val(tz['w'], tz['l']),
            pt['w'], pt['l'], pct_val(pt['w'], pt['l']),
            br['T']['w'], br['T']['l'], br['P']['w'], br['P']['l'],
            br['Z']['w'], br['Z']['l'],
        ] + [m['yearly'].get(y, 0) for y in ctx['years'][::-1]])
    counts['maps.csv'] = write_csv('maps.csv', [
        '맵', '총세트', '사용일수', '첫사용', '마지막사용',
        '저그승_vs프로토스', '저그패_vs프로토스', '저그승률_vs프로토스',
        '테란승_vs저그', '테란패_vs저그', '테란승률_vs저그',
        '프로토스승_vs테란', '프로토스패_vs테란', '프로토스승률_vs테란',
        '테란_세트승', '테란_세트패', '프로토스_세트승', '프로토스_세트패',
        '저그_세트승', '저그_세트패',
    ] + ['%s년_세트' % y for y in ctx['years'][::-1]], rows)

    rows = []
    for r in stats.rivalries(data['players']):
        rows.append([
            r['a'], render.RACE_LABEL[r['aRace']], r['b'], render.RACE_LABEL[r['bRace']],
            r['matches'], r['matchW'], r['matchL'],
            r['setW'], r['setL'], pct_val(r['setW'], r['setL']),
        ])
    counts['headtohead.csv'] = write_csv('headtohead.csv', [
        '선수A', '종족A', '선수B', '종족B', '맞대결수',
        'A매치승', 'A매치패', 'A세트승', 'A세트패', 'A세트승률',
    ], rows)

    rows = []
    for y in ctx['yearly']:
        mu = y['mu']
        rows.append([
            y['year'], y['matches'], y['sets'], y['players'],
            y['topPlayer'] or '', y['topWins'], y['topLosses'],
            mu['PvT']['w'], mu['PvT']['l'], pct_val(mu['PvT']['w'], mu['PvT']['l']),
            mu['TvZ']['w'], mu['TvZ']['l'], pct_val(mu['TvZ']['w'], mu['TvZ']['l']),
            mu['PvZ']['w'], mu['PvZ']['l'], pct_val(mu['PvZ']['w'], mu['PvZ']['l']),
        ])
    counts['yearly.csv'] = write_csv('yearly.csv', [
        '연도', '매치', '세트', '출전선수', '최다승선수', '최다승', '최다승선수패',
        '프로토스승_vs테란', '프로토스패_vs테란', '프로토스승률_vs테란',
        '테란승_vs저그', '테란패_vs저그', '테란승률_vs저그',
        '프로토스승_vs저그', '프로토스패_vs저그', '프로토스승률_vs저그',
    ], rows)
    return counts


def build_manifest(counts, built_at, built_ko):
    files = []
    for name, label, desc in CSV_SPEC:
        files.append({
            'label': label, 'desc': desc, 'file': name, 'group': '끝장전',
            'rows': counts.get(name, 0),
            'url': '%s/csv/%s' % (render.BASE_URL, name),
        })
    for name, label, desc in ASL_CSV_SPEC:
        files.append({
            'label': 'ASL ' + label, 'desc': desc, 'file': name, 'group': 'ASL',
            'rows': counts.get(name, 0),
            'url': '%s/csv/%s' % (render.BASE_URL, name),
        })
    manifest = {
        'updatedAt': built_at,
        'updatedAtKo': built_ko,
        'base': render.BASE_URL,
        'xlsx': '%s/xlsx/sc-endgame.xlsx' % render.BASE_URL,
        'aslXlsx': '%s/xlsx/asl.xlsx' % render.BASE_URL,
        'json': '%s/data/endgame.json' % render.BASE_URL,
        'aslJson': '%s/data/asl.json' % render.BASE_URL,
        'files': files,
    }
    write('csv/index.json', json.dumps(manifest, ensure_ascii=False, indent=2) + '\n')
    return manifest


# ── XLSX ───────────────────────────────────────────────────────
def build_xlsx(spec=None, out_name='sc-endgame.xlsx'):
    """CSV 여러 개를 시트 여러 장짜리 엑셀 한 파일로 묶습니다."""
    spec = spec or CSV_SPEC
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print('  ! openpyxl 이 없어 xlsx 는 건너뜁니다 (pip install openpyxl)')
        return None

    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color='FFFFFF', size=10)
    head_fill = PatternFill('solid', fgColor='1C8CFF')

    for name, label, _desc in spec:
        path = os.path.join(ROOT, 'csv', name)
        with io.open(path, encoding='utf-8-sig', newline='') as f:
            rows = list(csv.reader(f))
        ws = wb.create_sheet(label)
        for r in rows:
            ws.append([_num(c) for c in r])
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'
        if rows:
            ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(rows[0])), len(rows))
            for i in range(1, len(rows[0]) + 1):
                col = get_column_letter(i)
                width = max((_disp_len(r[i - 1]) for r in rows[:200] if len(r) >= i),
                            default=8)
                ws.column_dimensions[col].width = min(max(width + 2, 8), 42)

    out = os.path.join(ROOT, 'xlsx', out_name)
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    return os.path.getsize(out)


def _num(s):
    """숫자로 보이는 칸은 숫자로 넣습니다 — 엑셀에서 정렬·합계가 되도록."""
    if s == '':
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return s


def _disp_len(s):
    """한글은 대략 두 칸을 차지합니다."""
    return sum(2 if ord(c) > 0x2E80 else 1 for c in str(s))


# ── 메인 ───────────────────────────────────────────────────────
def apply_site_config():
    """사이트 주소를 설정 파일(또는 --base-url)에서 읽어 반영합니다."""
    cfg = read_json(os.path.join(ROOT, 'data', 'site.json'), {}) or {}
    base = cfg.get('baseUrl') or render.BASE_URL
    for i, a in enumerate(sys.argv):
        if a == '--base-url' and i + 1 < len(sys.argv):
            base = sys.argv[i + 1]
        elif a.startswith('--base-url='):
            base = a.split('=', 1)[1]
    render.BASE_URL = base.rstrip('/')

    cname = cfg.get('cname')
    path = os.path.join(ROOT, 'CNAME')
    if cname:
        write('CNAME', cname.strip() + '\n')
    elif os.path.exists(path):
        os.remove(path)          # 커스텀 도메인을 껐으면 파일도 치웁니다
    return render.BASE_URL, cname


def main():
    base_url, cname = apply_site_config()
    data, video_matched = load()
    ctx = enrich(data)

    now = datetime.now(timezone.utc)
    built_at = now.strftime('%Y-%m-%dT%H:%M:%S.') + '%03dZ' % (now.microsecond // 1000)
    built_ko = now.astimezone(KST).strftime('%Y년 %m월 %d일 %H:%M (KST)')
    ctx['builtAtKo'] = built_ko

    css = io.open(os.path.join(HERE, 'site.css'), encoding='utf-8').read()
    app_js = io.open(os.path.join(HERE, 'app.js'), encoding='utf-8').read()

    print('끝장전 기록실 빌드')
    print('  사이트 주소 %s%s' % (base_url, ('  · CNAME %s' % cname) if cname else ''))
    print('  매치 %d · 세트 %s · 선수 %d · 맵 %d · 연도 %s'
          % (data['global']['totalMatches'], format(data['global']['totalSets'], ','),
             len(data['players']), len(data['maps']),
             ctx['years'][-1] + '~' + ctx['years'][0]))

    asl = read_json(os.path.join(ROOT, 'data', 'asl.json'))
    asl_ctx = None
    if asl:
        asl_ctx = asl_enrich(asl)
        asl_ctx['builtAtKo'] = built_ko
        asl_ctx['endgameSlug'] = {p['name']: p['slug'] for p in data['players']}

    counts = build_csv(data, ctx)
    for name, _l, _d in CSV_SPEC:
        print('  csv/%-16s %5d행' % (name, counts[name]))

    size = build_xlsx()
    if size:
        print('  xlsx/sc-endgame.xlsx  %.0fKB' % (size / 1024))

    if asl:
        counts.update(asl_build_csv(asl, asl_ctx))
        for name, _l, _d in ASL_CSV_SPEC:
            print('  csv/%-16s %5d행' % (name, counts[name]))
        size = build_xlsx(ASL_CSV_SPEC, 'asl.xlsx')
        if size:
            print('  xlsx/asl.xlsx         %.0fKB' % (size / 1024))

    manifest = build_manifest(counts, built_at, built_ko)

    n = write('index.html', render.index_page(css, app_js, hub_data(data, ctx, built_at)))
    print('  index.html            %.0fKB' % (n / 1024))

    # 예전 파일이 남지 않도록 선수 페이지 폴더를 비우고 다시 만듭니다.
    pdir = os.path.join(ROOT, 'p')
    if os.path.isdir(pdir):
        for f in os.listdir(pdir):
            if f.endswith('.html'):
                os.remove(os.path.join(pdir, f))
    total = 0
    for p in data['players']:
        total += write('p/%s.html' % p['slug'], render.player_page(p, ctx, css))
    print('  p/*.html              %d개 · 합계 %.0fKB' % (len(data['players']), total / 1024))

    write('sheets.html', render.sheets_page(css, manifest))
    print('  sheets.html           작성')

    # CG 제작 툴 — 자동완성용 선수 목록은 두 기록실을 합쳐서 넣습니다.
    roster, seen = [], set()
    for src, tag in ((data['players'], '끝장전'), ((asl or {}).get('players', []), 'ASL')):
        for p in src:
            if p['name'] in seen:
                continue
            seen.add(p['name'])
            roster.append({'name': p['name'], 'race': p['race'], 'from': tag})
    roster.sort(key=lambda x: x['name'])
    cg_js = io.open(os.path.join(HERE, 'cg_app.js'), encoding='utf-8').read()
    n = write('admin/cg.php', render.cg_page(css, cg_js, roster))
    write('admin/auth.php', admin_php.AUTH_PHP)
    write('admin/logout.php', admin_php.LOGOUT_PHP)
    write('admin/.htaccess', admin_php.HTACCESS)
    write('admin/config.sample.php', admin_php.CONFIG_SAMPLE)
    write('admin/index.php', admin_php.index_php(css, render.SITE_NAME))
    # 예전 빌드가 남긴 무방비 파일이 있으면 치웁니다.
    old = os.path.join(ROOT, 'admin', 'cg.html')
    if os.path.exists(old):
        os.remove(old)
    print('  admin/*.php           %.0fKB · 선수 %d명 자동완성 · 로그인 보호'
          % (n / 1024, len(roster)))

    if asl:
        asl_js = io.open(os.path.join(HERE, 'asl_app.js'), encoding='utf-8').read()
        n = write('asl/index.html',
                  render.asl_index_page(css, asl_js, asl_hub_data(asl, asl_ctx, built_at)))
        print('  asl/index.html        %.0fKB' % (n / 1024))
        adir = os.path.join(ROOT, 'asl', 'p')
        if os.path.isdir(adir):
            for f in os.listdir(adir):
                if f.endswith('.html'):
                    os.remove(os.path.join(adir, f))
        total = 0
        for p in asl['players']:
            total += write('asl/p/%s.html' % p['slug'],
                           render.asl_player_page(p, asl_ctx, css))
        print('  asl/p/*.html          %d개 · 합계 %.0fKB'
              % (len(asl['players']), total / 1024))
        g = asl['global']
        print('    ASL 대회 %d · 매치 %s · 세트 %s · 선수 %d'
              % (g['totalTournaments'], format(g['totalMatches'], ','),
                 format(g['totalSets'], ','), g['totalPlayers']))
    else:
        print('  data/asl.json 이 없어 ASL 섹션은 건너뜁니다'
              ' (tools/asl_import.py 로 만드세요)')

    ok, msg = audit.report(ROOT, base_url)
    print('  ' + msg.replace('\n', '\n  '))

    if video_matched:
        print('  다시보기 영상 %d/%d 경기 연결' % (video_matched, len(data['matches'])))
    else:
        print('  다시보기 영상 연결 없음 — data/videos.json 이 비어 있습니다.')
        print('    (tools/fetch_videos.py 로 채우거나 직접 적어 넣으면 바로 반영됩니다)')
    print('완료.')



# ── ASL ────────────────────────────────────────────────────────
ASL_CSV_SPEC = [
    ('asl-players.csv', '선수', 'ASL 선수별 통산 성적과 우승 횟수'),
    ('asl-tournaments.csv', '대회', '대회별 규모와 우승·준우승'),
    ('asl-matches.csv', '매치', '매치(시리즈) 한 줄씩 — 대회·라운드·대진·스코어'),
    ('asl-sets.csv', '세트', '세트 한 줄씩 — 어떤 맵에서 누가 이겼는지'),
    ('asl-maps.csv', '맵', '맵별 사용 세트 수와 종족 상성'),
    ('asl-headtohead.csv', '상대전적', '선수 대 선수 맞대결 기록'),
]

ASL_MIN_SET = 100
ASL_MIN_MATCH = 30


def asl_short(name):
    m = re.search(r'Season\s*(\d+)', name)
    return 'S%d' % int(m.group(1)) if m else name.split()[0]


def asl_rivalries(players, limit=None):
    seen, rows = set(), []
    race_of = {p['name']: p['race'] for p in players}
    for p in players:
        for o in p['vsPlayers']:
            if o['name'] not in race_of:
                continue
            pair = tuple(sorted([p['name'], o['name']]))
            if pair in seen:
                continue
            seen.add(pair)
            a, b = pair
            flip = p['name'] != a
            setW, setL = (o['l'], o['w']) if flip else (o['w'], o['l'])
            mW, mL = (o['ml'], o['mw']) if flip else (o['mw'], o['ml'])
            rows.append({'a': a, 'aRace': race_of[a], 'b': b, 'bRace': race_of[b],
                         'setW': setW, 'setL': setL, 'matchW': mW, 'matchL': mL,
                         'matches': mW + mL})
    rows.sort(key=lambda r: (-(r['setW'] + r['setL']), -r['matches'], r['a']))
    return rows[:limit] if limit else rows


def asl_enrich(data):
    """슬러그·최고 라운드·기록실 순위를 붙입니다."""
    names = [p['name'] for p in data['players']]
    slugs = unique_slugs(names)
    race_of = {p['name']: p['race'] for p in data['players']}
    order = data['roundOrder']
    rank = {r: i for i, r in enumerate(order)}

    for t in data['tournaments']:
        t['short'] = asl_short(t['name'])

    for p in data['players']:
        p['slug'] = slugs[p['name']]
        best, best_i = '', -1
        for v in p['byTour'].values():
            if rank.get(v.get('best', ''), -1) > best_i:
                best_i, best = rank.get(v.get('best', ''), -1), v.get('best', '')
        p['bestRound'] = best

    def prow(p, value, label, detail):
        return {'name': p['name'], 'race': p['race'], 'slug': p['slug'],
                'value': value, 'label': label, 'detail': detail}

    def top(rows, n=10):
        return sorted(rows, key=lambda r: -r['value'])[:n]

    P = data['players']
    records = {
        'titles': top([prow(p, p['titles'], '%d회' % p['titles'],
                            '준우승 %d회 · %d개 대회 출전' % (p['runnerUps'], p['tournaments']))
                       for p in P if p['titles']]),
        'setWin': top([prow(p, p['setWin'], '%d승' % p['setWin'],
                            '%d승 %d패 · %s' % (p['setWin'], p['setLoss'],
                                              render.pct(p['setWin'], p['setLoss'])))
                       for p in P]),
        'setPct': top([prow(p, pct_val(p['setWin'], p['setLoss']) or 0,
                            '%s' % render.pct(p['setWin'], p['setLoss']),
                            '%d승 %d패' % (p['setWin'], p['setLoss']))
                       for p in P if p['setWin'] + p['setLoss'] >= ASL_MIN_SET]),
        'matchWin': top([prow(p, p['matchWin'], '%d승' % p['matchWin'],
                              '%d승 %d패 · %s' % (p['matchWin'], p['matchLoss'],
                                                render.pct(p['matchWin'], p['matchLoss'])))
                         for p in P]),
        'matchPct': top([prow(p, pct_val(p['matchWin'], p['matchLoss']) or 0,
                              '%s' % render.pct(p['matchWin'], p['matchLoss']),
                              '%d승 %d패' % (p['matchWin'], p['matchLoss']))
                         for p in P if p['matchWin'] + p['matchLoss'] >= ASL_MIN_MATCH]),
        'tournaments': top([prow(p, p['tournaments'], '%d개' % p['tournaments'],
                                 '최고 성적 %s' % (p['bestRound'] or '-')) for p in P]),
        'minSet': ASL_MIN_SET, 'minMatch': ASL_MIN_MATCH,
    }

    return {
        'slugs': slugs, 'raceOf': race_of,
        'tournaments': data['tournaments'], 'matches': data['matches'],
        'records': records,
        'rivalries': asl_rivalries(data['players'], 15),
        'champOf': {t['name']: t['champion'] for t in data['tournaments'] if t['champion']},
        'endgameSlug': {},
    }


def asl_global_mu(data):
    mu = stats.blank_mu()
    for m in data['maps']:
        for k in mu:
            mu[k]['w'] += m['mu'][k]['w']
            mu[k]['l'] += m['mu'][k]['l']
    return mu


def asl_hub_data(data, ctx, built_at):
    players = [{
        'name': p['name'], 'slug': p['slug'], 'race': p['race'],
        'setWin': p['setWin'], 'setLoss': p['setLoss'],
        'matchWin': p['matchWin'], 'matchLoss': p['matchLoss'],
        'titles': p['titles'], 'runnerUps': p['runnerUps'],
        'tournaments': p['tournaments'], 'bestRound': p['bestRound'],
        'byTour': p['byTour'],
    } for p in data['players']]
    maps = [{'name': m['name'], 'totalSets': m['totalSets'], 'mu': m['mu'],
             'mirrorSets': m.get('mirrorSets', 0),
             'byTour': m['byTour'], 'tournaments': m['tournaments']}
            for m in data['maps']]
    return {
        'builtAt': built_at,
        'global': data['global'],
        'tournaments': data['tournaments'],
        'roundOrder': data['roundOrder'],
        'mu': asl_global_mu(data),
        'records': ctx['records'],
        'rivalries': ctx['rivalries'],
        'champOf': ctx['champOf'],
        'slugs': ctx['slugs'],
        'players': players,
        'maps': maps,
        'matches': data['matches'],
    }


def asl_build_csv(data, ctx):
    counts = {}
    rows = []
    for p in sorted(data['players'], key=lambda x: (-x['titles'], -x['setWin'], x['name'])):
        v = p['vsRace']
        rows.append([
            p['name'], p['slug'], render.RACE_LABEL[p['race']], p['race'],
            p['titles'], p['runnerUps'], p['tournaments'], p['bestRound'],
            p['matchWin'], p['matchLoss'], pct_val(p['matchWin'], p['matchLoss']),
            p['setWin'], p['setLoss'], pct_val(p['setWin'], p['setLoss']),
            v['T']['w'], v['T']['l'], v['P']['w'], v['P']['l'], v['Z']['w'], v['Z']['l'],
        ])
    counts['asl-players.csv'] = write_csv('asl-players.csv', [
        '선수', '슬러그', '종족', '종족코드', '우승', '준우승', '출전대회', '최고성적',
        '매치승', '매치패', '매치승률', '세트승', '세트패', '세트승률',
        'vs테란_세트승', 'vs테란_세트패', 'vs프로토스_세트승', 'vs프로토스_세트패',
        'vs저그_세트승', 'vs저그_세트패',
    ], rows)

    rows = []
    for t in data['tournaments']:
        mu = t['mu']
        rows.append([
            t['name'], t['matches'], t['sets'], t['players'],
            len(t['rounds']), ' | '.join(r['name'] for r in t['rounds']),
            t.get('mirrorSets', 0),
            t['champion'] or '', t['championRace'] or '',
            t['finalScore'] or '', t['runnerUp'] or '', t['runnerUpRace'] or '',
            mu['PvT']['w'], mu['PvT']['l'], mu['TvZ']['w'], mu['TvZ']['l'],
            mu['PvZ']['w'], mu['PvZ']['l'],
        ])
    counts['asl-tournaments.csv'] = write_csv('asl-tournaments.csv', [
        '대회', '매치', '세트', '선수', '라운드수', '라운드', '동족전세트',
        '우승', '우승종족', '결승스코어', '준우승', '준우승종족',
        '프로토스승_vs테란', '프로토스패_vs테란', '테란승_vs저그', '테란패_vs저그',
        '프로토스승_vs저그', '프로토스패_vs저그',
    ], rows)

    rows, set_rows = [], []
    for m in data['matches']:
        a, b = m['players']
        rows.append([
            m['tournament'], m['round'], a, m['race'][a], b, m['race'][b],
            m['setWins'][a], m['setWins'][b], m['winner'] or '', m['sets'],
            ' | '.join(x for x in m['maps'] if x),
        ])
        for i, name in enumerate(m['maps'], 1):
            set_rows.append([m['tournament'], m['round'], a, b, i, name])
    counts['asl-matches.csv'] = write_csv('asl-matches.csv', [
        '대회', '라운드', '선수A', '종족A', '선수B', '종족B',
        '세트A', '세트B', '승자', '총세트', '맵목록',
    ], rows)
    counts['asl-sets.csv'] = write_csv('asl-sets.csv',
                                       ['대회', '라운드', '선수A', '선수B', '세트번호', '맵'],
                                       set_rows)

    rows = []
    for m in data['maps']:
        mu, br = m['mu'], m['byRace']
        rows.append([
            m['name'], m['totalSets'], m['tournaments'], m.get('mirrorSets', 0),
            mu['PvT']['w'], mu['PvT']['l'], pct_val(mu['PvT']['w'], mu['PvT']['l']),
            mu['TvZ']['w'], mu['TvZ']['l'], pct_val(mu['TvZ']['w'], mu['TvZ']['l']),
            mu['PvZ']['w'], mu['PvZ']['l'], pct_val(mu['PvZ']['w'], mu['PvZ']['l']),
            br['T']['w'], br['T']['l'], br['P']['w'], br['P']['l'],
            br['Z']['w'], br['Z']['l'],
        ])
    counts['asl-maps.csv'] = write_csv('asl-maps.csv', [
        '맵', '총세트', '사용대회', '동족전세트',
        '프로토스승_vs테란', '프로토스패_vs테란', '프로토스승률_vs테란',
        '테란승_vs저그', '테란패_vs저그', '테란승률_vs저그',
        '프로토스승_vs저그', '프로토스패_vs저그', '프로토스승률_vs저그',
        '테란_세트승', '테란_세트패', '프로토스_세트승', '프로토스_세트패',
        '저그_세트승', '저그_세트패',
    ], rows)

    rows = []
    for r in asl_rivalries(data['players']):
        rows.append([
            r['a'], render.RACE_LABEL[r['aRace']], r['b'], render.RACE_LABEL[r['bRace']],
            r['matches'], r['matchW'], r['matchL'],
            r['setW'], r['setL'], pct_val(r['setW'], r['setL']),
        ])
    counts['asl-headtohead.csv'] = write_csv('asl-headtohead.csv', [
        '선수A', '종족A', '선수B', '종족B', '맞대결매치',
        'A매치승', 'A매치패', 'A세트승', 'A세트패', 'A세트승률',
    ], rows)
    return counts

if __name__ == '__main__':
    main()
