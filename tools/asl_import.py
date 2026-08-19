#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ASL 엑셀(세트 단위 기록)을 data/asl.json 정본으로 옮깁니다.

    python3 tools/asl_import.py <ASL.xlsx>

엑셀 한 줄이 세트 하나입니다.
  경기번호 | 구분 | 선수명A | 종족A | 승/패A | MAP | 승/패B | 종족B | 선수명B

'구분' 은 병합된 칸이라 첫 줄에만 값이 있어 아래로 채워 씁니다.
'구분' 은 "ASL Season 21\\n8강" 처럼 대회와 라운드가 줄바꿈으로 붙어 있습니다.

매치(시리즈)는 같은 라운드 안에서 같은 두 선수가 연달아 나오는 구간으로 묶습니다.
조별 리그처럼 같은 두 사람이 라운드 안에서 떨어져 두 번 만나면 별개 매치가 됩니다.
"""

import argparse
import io
import json
import os
import re
import sys
from collections import Counter, defaultdict, OrderedDict

# 윈도우 콘솔에서 한글·기호가 깨지거나 터지지 않게 합니다.
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
except (AttributeError, OSError):
    pass


HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

RACES = ['T', 'P', 'Z']
MATCHUPS = [('P', 'T'), ('T', 'Z'), ('P', 'Z')]
# 라운드는 대회 진행 순서대로 놓습니다.
ROUND_ORDER = ['와일드카드', '32강', '24강', '16강', '8강', '4강', '3-4위전', '결승전']
FINAL_NAMES = {'결승전'}
# 시트에 같은 라운드가 여러 이름으로 적혀 있어 하나로 맞춥니다.
ROUND_ALIAS = {'WILD CARD': '와일드카드', '결승': '결승전', 'FINALS': '결승전'}
TOURNEY_RE = re.compile(r'^(대국민\s*스타리그|ASL\s*Season\s*\d+)\s*(.*)$')


def norm(s):
    """줄바꿈·겹공백을 한 칸으로 눌러 줍니다."""
    return ' '.join(str(s or '').split())


def split_group(text):
    """'ASL Season 21 8강' → ('ASL Season 21', '8강'). 'WILD CARD' 처럼
    낱말이 띄어져 있어도 대회명 쪽을 먼저 떼어 내 안전하게 가릅니다."""
    t = norm(text)
    m = TOURNEY_RE.match(t)
    if m:
        tour, rnd = norm(m.group(1)), norm(m.group(2))
    else:
        for r in sorted(ROUND_ORDER, key=len, reverse=True):
            if t.endswith(' ' + r) or t == r:
                tour, rnd = t[:-len(r)].strip(), r
                break
        else:
            mm = re.search(r'(\S+)$', t)
            tour, rnd = (t[:mm.start()].strip(), mm.group(1)) if mm else (t, '')
    return tour, ROUND_ALIAS.get(rnd, rnd)


def tourney_sort_key(name):
    m = re.search(r'Season\s*(\d+)', name)
    return (1, int(m.group(1))) if m else (0, 0)


def tourney_id(name):
    m = re.search(r'Season\s*(\d+)', name)
    return 'asl%02d' % int(m.group(1)) if m else 'pre'


def mu_key(r1, r2):
    if r1 == r2:
        return None
    for a, b in MATCHUPS:
        if {r1, r2} == {a, b}:
            return a + 'v' + b
    return None


def blank_mu():
    return {a + 'v' + b: {'w': 0, 'l': 0} for a, b in MATCHUPS}


def blank_mirror():
    return {r: 0 for r in RACES}


def add_mu(mu, race_win, race_lose):
    k = mu_key(race_win, race_lose)
    if not k:
        return
    if k[0] == race_win:
        mu[k]['w'] += 1
    else:
        mu[k]['l'] += 1


def rows_to_sets(rows):
    """시트 3행부터의 줄들을 세트 목록으로. 엑셀과 CSV 가 함께 씁니다.

    한 줄 = (경기번호, 구분, 선수명A, 종족A, 승/패A, MAP, 승/패B, 종족B, 선수명B)
    '구분' 은 병합된 칸이라 첫 줄에만 값이 있어 아래로 채워 씁니다.
    """
    out, group = [], None
    for row in rows:
        cells = list(row[:9]) + [None] * max(0, 9 - len(row))
        num, grp, a, ra, wa, mp, wbb, rb, b = cells[:9]
        if grp not in (None, ''):
            group = grp
        if not (a and b and group):
            continue
        a, b = norm(a), norm(b)
        ra, rb = norm(ra).upper(), norm(rb).upper()
        wa, wbb = norm(wa), norm(wbb)
        if ra not in RACES or rb not in RACES or {wa, wbb} != {'승', '패'}:
            continue
        tour, rnd = split_group(group)
        winner, loser = (a, b) if wa == '승' else (b, a)
        out.append({
            'no': int(num) if isinstance(num, (int, float)) else len(out) + 1,
            'tournament': tour, 'round': rnd,
            'a': a, 'aRace': ra, 'b': b, 'bRace': rb,
            'map': norm(mp), 'winner': winner, 'loser': loser,
            'winRace': ra if winner == a else rb,
            'loseRace': rb if winner == a else ra,
        })
    return out


def read_sets(path):
    """엑셀 파일에서 읽습니다."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return rows_to_sets(ws.iter_rows(min_row=3, values_only=True))


def read_sets_csv(text):
    """구글시트 CSV 내보내기에서 읽습니다. 병합 칸은 빈칸으로 나옵니다."""
    import csv as _csv
    rows = list(_csv.reader(io.StringIO(text)))
    return rows_to_sets(rows[2:])          # 1행 제목, 2행 머리글


def fetch_sheet_csv(source=None):
    """구글시트를 CSV 로 받아옵니다. 시트가 링크 공개면 인증이 필요 없습니다."""
    import urllib.error
    import urllib.request
    src = source or {}
    if not src:
        path = os.path.join(ROOT, 'data', 'asl-source.json')
        if not os.path.exists(path):
            raise SystemExit('data/asl-source.json 이 없습니다. sheetId 와 gid 를 넣어 주세요.')
        src = json.load(io.open(path, encoding='utf-8'))
    sid = src.get('sheetId')
    gid = src.get('gid')
    gid = '' if gid in (None, '') else str(gid)
    if not sid:
        raise SystemExit('data/asl-source.json 에 sheetId 가 없습니다.')
    # gid 를 비워 두면 첫 번째 탭을 받아옵니다. 새로 만든 시트는 첫 탭 gid 가
    # 0 이 아닐 수 있어서, 탭이 하나뿐이면 그냥 비워 두는 편이 안전합니다.
    url = ('https://docs.google.com/spreadsheets/d/%s/export?format=csv' % sid
           + ('&gid=%s' % gid if gid else ''))
    print('구글시트에서 받아오는 중 — %s (%s)'
          % (src.get('title') or sid, ('gid %s' % gid) if gid else '첫 번째 탭'))
    req = urllib.request.Request(url, headers={'User-Agent': 'sc-endgame/1.0'})
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            raw = r.read()
    except urllib.error.HTTPError as e:
        raise SystemExit(
            '시트를 받지 못했습니다 (HTTP %s).\n'
            "  시트 공유 설정이 '링크가 있는 모든 사용자 — 뷰어' 인지 확인해 주세요.\n"
            '  주소: %s' % (e.code, url))
    except urllib.error.URLError as e:
        raise SystemExit('시트에 접속하지 못했습니다: %s' % e.reason)
    text = raw.decode('utf-8-sig')
    if text.lstrip().startswith('<'):
        raise SystemExit(
            'CSV 대신 로그인 화면이 왔습니다.\n'
            "  시트를 '링크가 있는 모든 사용자 — 뷰어' 로 열어 주세요.")
    return text


def fix_races(sets):
    """한 선수의 종족은 하나입니다.

    시트에는 드물게 오타로 같은 선수가 다른 종족으로 적힌 줄이 있습니다.
    그대로 두면 상성·동족전 집계가 틀어지므로, 가장 많이 적힌 종족으로 맞추고
    무엇을 바꿨는지 돌려줍니다 (원본 시트는 건드리지 않습니다).
    """
    tally = defaultdict(Counter)
    for s in sets:
        tally[s['a']][s['aRace']] += 1
        tally[s['b']][s['bRace']] += 1

    canon, notes = {}, []
    for name, c in tally.items():
        best = c.most_common(1)[0][0]
        canon[name] = best
        if len(c) > 1:
            notes.append((name, best, {k: v for k, v in c.items() if k != best}))

    for s in sets:
        s['aRace'], s['bRace'] = canon[s['a']], canon[s['b']]
        s['winRace'] = canon[s['winner']]
        s['loseRace'] = canon[s['loser']]
    return notes


def load_sets(xlsx=None):
    """엑셀이나 구글시트에서 읽고 종족까지 맞춰서 돌려줍니다.

    어느 경로로 들어와도 같은 결과가 나오도록 여기 한 군데로 모읍니다.
    돌려주는 값은 (세트 목록, 종족을 고친 내역) 입니다.
    """
    sets = read_sets(xlsx) if xlsx else read_sets_csv(fetch_sheet_csv())
    return sets, fix_races(sets)


def group_matches(sets):
    """같은 라운드에서 같은 두 선수가 연달아 나오는 구간 = 매치 하나."""
    matches, cur = [], None
    for s in sets:
        pair = tuple(sorted([s['a'], s['b']]))
        sig = (s['tournament'], s['round'], pair)
        if cur is None or cur['_sig'] != sig:
            cur = {
                '_sig': sig, 'tournament': s['tournament'], 'round': s['round'],
                'players': list(pair),
                'race': {s['a']: s['aRace'], s['b']: s['bRace']},
                'setWins': {pair[0]: 0, pair[1]: 0},
                'maps': [], 'firstNo': s['no'],
            }
            matches.append(cur)
        cur['setWins'][s['winner']] += 1
        cur['maps'].append(s['map'])
        cur['race'].setdefault(s['a'], s['aRace'])
        cur['race'].setdefault(s['b'], s['bRace'])
    for m in matches:
        del m['_sig']
        x, y = m['players']
        m['winner'] = x if m['setWins'][x] > m['setWins'][y] else (
            y if m['setWins'][y] > m['setWins'][x] else None)
        m['sets'] = m['setWins'][x] + m['setWins'][y]
    return matches


def build(sets, matches):
    tours = OrderedDict()
    for s in sets:
        t = tours.setdefault(s['tournament'], {
            'id': tourney_id(s['tournament']), 'name': s['tournament'],
            'rounds': OrderedDict(), 'sets': 0, 'players': set(),
            'mu': blank_mu(), 'mirror': blank_mirror(),
        })
        r = t['rounds'].setdefault(s['round'], {'name': s['round'], 'sets': 0, 'matches': 0})
        r['sets'] += 1
        t['sets'] += 1
        t['players'].update([s['a'], s['b']])
        add_mu(t['mu'], s['winRace'], s['loseRace'])
        if s['winRace'] == s['loseRace']:
            t['mirror'][s['winRace']] += 1        # 동족전은 상성 집계에서 빠집니다

    for m in matches:
        t = tours[m['tournament']]
        t['rounds'][m['round']]['matches'] += 1

    # 우승·준우승은 결승 시리즈의 승자·패자입니다.
    for m in matches:
        if m['round'] in FINAL_NAMES and m['winner']:
            t = tours[m['tournament']]
            t['champion'] = m['winner']
            t['championRace'] = m['race'][m['winner']]
            other = [p for p in m['players'] if p != m['winner']][0]
            t['runnerUp'] = other
            t['runnerUpRace'] = m['race'][other]
            t['finalScore'] = '%d-%d' % (m['setWins'][m['winner']], m['setWins'][other])

    tour_list = []
    for name in sorted(tours, key=lambda n: tourney_sort_key(n)):
        t = tours[name]
        rounds = [t['rounds'][r] for r in sorted(
            t['rounds'], key=lambda r: (ROUND_ORDER.index(r)
                                        if r in ROUND_ORDER else len(ROUND_ORDER)))]
        tour_list.append({
            'id': t['id'], 'name': t['name'], 'rounds': rounds,
            'sets': t['sets'], 'matches': sum(r['matches'] for r in rounds),
            'players': len(t['players']), 'mu': t['mu'], 'mirror': t['mirror'],
            'mirrorSets': sum(t['mirror'].values()),
            'champion': t.get('champion'), 'championRace': t.get('championRace'),
            'runnerUp': t.get('runnerUp'), 'runnerUpRace': t.get('runnerUpRace'),
            'finalScore': t.get('finalScore'),
        })
    tour_list.reverse()                        # 최신 시즌이 위로

    # ── 선수 ──────────────────────────────────────────
    P = {}

    def slot(name, race):
        p = P.setdefault(name, {
            'name': name, 'race': race, 'setWin': 0, 'setLoss': 0,
            'matchWin': 0, 'matchLoss': 0, 'titles': 0, 'runnerUps': 0,
            'vsRace': {r: {'w': 0, 'l': 0} for r in RACES},
            'vs': defaultdict(lambda: {'w': 0, 'l': 0, 'mw': 0, 'ml': 0}),
            'byTour': {}, 'rounds': Counter(),
        })
        return p

    for s in sets:
        pw, pl = slot(s['winner'], s['winRace']), slot(s['loser'], s['loseRace'])
        pw['setWin'] += 1
        pl['setLoss'] += 1
        pw['vsRace'][s['loseRace']]['w'] += 1
        pl['vsRace'][s['winRace']]['l'] += 1
        pw['vs'][s['loser']]['w'] += 1
        pl['vs'][s['winner']]['l'] += 1
        for p, tag in ((pw, 'w'), (pl, 'l')):
            bt = p['byTour'].setdefault(s['tournament'], {
                'setWin': 0, 'setLoss': 0, 'matchWin': 0, 'matchLoss': 0, 'best': ''})
            bt['setWin' if tag == 'w' else 'setLoss'] += 1

    for m in matches:
        x, y = m['players']
        if not m['winner']:
            continue
        loser = y if m['winner'] == x else x
        pw, pl = P[m['winner']], P[loser]
        pw['matchWin'] += 1
        pl['matchLoss'] += 1
        pw['vs'][loser]['mw'] += 1
        pl['vs'][m['winner']]['ml'] += 1
        pw['byTour'][m['tournament']]['matchWin'] += 1
        pl['byTour'][m['tournament']]['matchLoss'] += 1
        for p in (pw, pl):
            p['rounds'][m['round']] += 1

    for t in tour_list:
        if t['champion']:
            P[t['champion']]['titles'] += 1
        if t['runnerUp']:
            P[t['runnerUp']]['runnerUps'] += 1

    # 대회별 최고 라운드
    rank = {r: i for i, r in enumerate(ROUND_ORDER)}
    for s in sets:
        for name in (s['a'], s['b']):
            bt = P[name]['byTour'][s['tournament']]
            if rank.get(s['round'], -1) > rank.get(bt['best'], -1):
                bt['best'] = s['round']

    players = []
    for name, p in P.items():
        vs = [{'name': k, 'race': P[k]['race'] if k in P else '',
               'w': v['w'], 'l': v['l'], 'mw': v['mw'], 'ml': v['ml']}
              for k, v in p['vs'].items()]
        vs.sort(key=lambda x: (-(x['w'] + x['l']), x['name']))
        players.append({
            'name': name, 'race': p['race'],
            'setWin': p['setWin'], 'setLoss': p['setLoss'],
            'matchWin': p['matchWin'], 'matchLoss': p['matchLoss'],
            'titles': p['titles'], 'runnerUps': p['runnerUps'],
            'tournaments': len(p['byTour']),
            'vsRace': p['vsRace'], 'vsPlayers': vs, 'byTour': p['byTour'],
        })
    players.sort(key=lambda p: (-p['titles'], -p['setWin'], p['name']))

    # ── 맵 ────────────────────────────────────────────
    M = {}
    for s in sets:
        if not s['map']:
            continue
        m = M.setdefault(s['map'], {
            'name': s['map'], 'totalSets': 0, 'mu': blank_mu(),
            'mirror': blank_mirror(), 'mirrorSets': 0,
            'byRace': {r: {'w': 0, 'l': 0} for r in RACES}, 'byTour': Counter(),
        })
        m['totalSets'] += 1
        add_mu(m['mu'], s['winRace'], s['loseRace'])
        if s['winRace'] == s['loseRace']:
            m['mirror'][s['winRace']] += 1
            m['mirrorSets'] += 1
        m['byRace'][s['winRace']]['w'] += 1
        m['byRace'][s['loseRace']]['l'] += 1
        m['byTour'][s['tournament']] += 1
    maps = []
    for m in M.values():
        m['byTour'] = dict(m['byTour'])
        m['tournaments'] = len(m['byTour'])
        maps.append(m)
    maps.sort(key=lambda m: -m['totalSets'])

    blank_maps = sum(1 for s in sets if not s['map'])
    mirror_total = blank_mirror()
    for s in sets:
        if s['winRace'] == s['loseRace']:
            mirror_total[s['winRace']] += 1

    return {
        'global': {
            'totalSets': len(sets), 'totalMatches': len(matches),
            'totalPlayers': len(players), 'totalMaps': len(maps),
            'totalTournaments': len(tour_list),
            'firstTournament': tour_list[-1]['name'], 'lastTournament': tour_list[0]['name'],
            'setsWithoutMap': blank_maps,
            'mirror': mirror_total,
            'mirrorSets': sum(mirror_total.values()),
        },
        'tournaments': tour_list,
        'players': players,
        'maps': maps,
        'matches': [{
            'tournament': m['tournament'], 'round': m['round'], 'no': m['firstNo'],
            'players': m['players'], 'race': m['race'], 'setWins': m['setWins'],
            'winner': m['winner'], 'sets': m['sets'], 'maps': m['maps'],
        } for m in matches],
        'roundOrder': ROUND_ORDER,
        'setList': set_list(sets, players, maps, tour_list),
    }


def set_list(sets, players, maps, tour_list):
    """세트 하나하나를 숫자 배열로 압축합니다 — 상대전적 조회에 씁니다.

    matches 는 세트 수만 갖고 있어서 '누가 어느 맵에서 이겼는지' 를 모릅니다.
    선수 vs 선수, 종족 vs 종족, 맵별 전적을 내려면 세트 단위가 필요합니다.

    rows 한 줄 = [a, b, map, tour, round, aRace, bRace, winner]
      a·b·map·tour  각 배열에서의 자리 번호 (map 은 없으면 -1)
      round         rounds 배열에서의 자리 번호
      aRace·bRace   0=T 1=P 2=Z
      winner        0 이면 a 가, 1 이면 b 가 이김
    """
    pi = {p['name']: i for i, p in enumerate(players)}
    mi = {m['name']: i for i, m in enumerate(maps)}
    ti = {t['name']: i for i, t in enumerate(tour_list)}
    rounds, ri = [], {}
    ri_get = ri.setdefault
    rows = []
    for s in sets:
        rnd = s['round']
        if rnd not in ri:
            ri_get(rnd, len(rounds))
            rounds.append(rnd)
        rows.append([
            pi[s['a']], pi[s['b']],
            mi.get(s['map'], -1) if s['map'] else -1,
            ti[s['tournament']], ri[rnd],
            RACES.index(s['aRace']), RACES.index(s['bRace']),
            0 if s['winner'] == s['a'] else 1,
        ])
    return {'races': list(RACES), 'rounds': rounds, 'rows': rows}


def summarize(data):
    """대회 → 라운드별 세트 수. 두 판을 견주는 데 씁니다."""
    out = {}
    for t in data['tournaments']:
        out[t['name']] = {r['name']: r['sets'] for r in t['rounds']}
    return out


def show_diff(old, new):
    """지금 data/asl.json 과 새로 읽은 시트가 어떻게 다른지 보여 줍니다."""
    if not old:
        print('  (data/asl.json 이 없어 전부 새 기록입니다)')
        return True

    o, n = summarize(old), summarize(new)
    added_t = [t for t in n if t not in o]
    gone_t = [t for t in o if t not in n]
    changed = []
    for t in n:
        if t not in o:
            continue
        for r, c in n[t].items():
            before = o[t].get(r)
            if before is None:
                changed.append(('새 라운드', t, r, 0, c))
            elif c > before:
                changed.append(('경기 추가', t, r, before, c))
            elif c < before:
                changed.append(('세트가 줄었음', t, r, before, c))
        for r, c in o[t].items():
            if r not in n[t]:
                changed.append(('라운드 사라짐', t, r, c, 0))

    op = {p['name'] for p in old['players']}
    np_ = {p['name'] for p in new['players']}

    print('  세트 %d → %d (%+d) · 매치 %d → %d (%+d)'
          % (old['global']['totalSets'], new['global']['totalSets'],
             new['global']['totalSets'] - old['global']['totalSets'],
             old['global']['totalMatches'], new['global']['totalMatches'],
             new['global']['totalMatches'] - old['global']['totalMatches']))
    if added_t:
        print('  새 대회: ' + ', '.join(added_t))
    if np_ - op:
        print('  새 선수: ' + ', '.join(sorted(np_ - op)))
    for kind, t, r, a, b in changed:
        ok = kind in ('새 라운드', '경기 추가')          # 대회가 진행되면 자연스러운 변화
        print('%s %s — %s %s (%d → %d세트)'
              % ('  +' if ok else '  !', t, r, kind, a, b))
    if gone_t:
        print('  ! 사라진 대회: ' + ', '.join(gone_t) + '  — 시트를 확인해 주세요')

    risky = [c for c in changed
             if c[0] not in ('새 라운드', '경기 추가')] or gone_t
    if risky:
        print('\n  ! 기록이 줄거나 사라졌습니다. 시트에서 지워진 줄이 없는지 확인하세요.')
    elif not (added_t or changed):
        print('  달라진 것이 없습니다.')
    return True


def main():
    ap = argparse.ArgumentParser(
        description='ASL 기록을 구글시트나 엑셀에서 읽어 data/asl.json 으로 옮깁니다.')
    ap.add_argument('xlsx', nargs='?',
                    help='ASL 엑셀 파일 (생략하고 --sheet 를 쓰면 구글시트에서 받습니다)')
    ap.add_argument('--sheet', action='store_true',
                    help='data/asl-source.json 의 구글시트에서 바로 받아옵니다')
    ap.add_argument('--diff', action='store_true',
                    help='저장하지 않고 지금 데이터와 무엇이 다른지만 봅니다')
    args = ap.parse_args()

    if not args.sheet and not args.xlsx:
        print('(엑셀 파일을 주지 않아 구글시트에서 받아옵니다)')
    sets, fixes = load_sets(None if args.sheet else args.xlsx)
    matches = group_matches(sets)
    data = build(sets, matches)

    out = os.path.join(ROOT, 'data', 'asl.json')
    old = None
    if os.path.exists(out):
        try:
            old = json.load(io.open(out, encoding='utf-8'))
        except ValueError:
            old = None

    print('지금 데이터와 견주기')
    show_diff(old, data)
    print()

    if args.diff:
        print('--diff 라서 저장하지 않았습니다.')
        return

    io.open(out, 'w', encoding='utf-8').write(
        json.dumps(data, ensure_ascii=False, indent=1) + '\n')

    g = data['global']
    print('ASL 데이터 정리 완료 → data/asl.json')
    print('  대회 %d개 (%s ~ %s)' % (g['totalTournaments'],
                                  g['firstTournament'], g['lastTournament']))
    print('  세트 %d · 매치 %d · 선수 %d · 맵 %d'
          % (g['totalSets'], g['totalMatches'], g['totalPlayers'], g['totalMaps']))
    if g['setsWithoutMap']:
        print('  맵 이름이 비어 있는 세트 %d개' % g['setsWithoutMap'])
    print('  동족전 %d세트 (T %d · P %d · Z %d) — 상성 집계에서는 빠집니다'
          % (g['mirrorSets'], g['mirror']['T'], g['mirror']['P'], g['mirror']['Z']))
    champs = [t for t in data['tournaments'] if t['champion']]
    print('  우승 기록이 있는 대회 %d개 / 결승 미확정 %d개'
          % (len(champs), g['totalTournaments'] - len(champs)))
    if fixes:
        print('\n  ! 시트에 종족이 엇갈리게 적힌 선수가 있어 최빈 종족으로 맞췄습니다.')
        print('    (원본 엑셀은 그대로입니다 — 시트를 고치시면 이 줄이 사라집니다)')
        for name, best, others in fixes:
            print('      %s → %s 로 통일 (다르게 적힌 줄: %s)'
                  % (name, best,
                     ', '.join('%s %d줄' % (k, v) for k, v in sorted(others.items()))))


if __name__ == '__main__':
    main()
